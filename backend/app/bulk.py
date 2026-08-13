"""Helpers to parse uploaded CSV / XLSX files into row dictionaries."""
from __future__ import annotations

import csv
import io

from fastapi import HTTPException, UploadFile, status

MAX_UPLOAD_BYTES = 2 * 1024 * 1024  # 2 MB is plenty for a shop's customer/item list
MAX_ROWS = 5000


def _norm_header(h: object) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def parse_upload(file: UploadFile, content: bytes) -> list[dict[str, str]]:
    """Return a list of {header: value} dicts from a CSV or XLSX upload.

    Headers are normalised to lowercase_with_underscores. Values are strings
    (empty string for blanks). Raises HTTPException on malformed input.
    """
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="File too large (max 2 MB).",
        )

    name = (file.filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        rows = _parse_xlsx(content)
    elif name.endswith(".csv") or (file.content_type or "").startswith("text/"):
        rows = _parse_csv(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload a .csv or .xlsx file.",
        )

    if len(rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Too many rows (max {MAX_ROWS}).",
        )
    return rows


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        return []
    field_map = {fn: _norm_header(fn) for fn in reader.fieldnames}
    out: list[dict[str, str]] = []
    for raw in reader:
        out.append({field_map[k]: (str(v).strip() if v is not None else "") for k, v in raw.items() if k in field_map})
    return out


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read the Excel file. Is it a valid .xlsx?",
        )
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [_norm_header(h) for h in header_row]
    out: list[dict[str, str]] = []
    for row in rows_iter:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue  # skip fully blank rows
        record: dict[str, str] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            value = row[idx] if idx < len(row) else None
            record[h] = "" if value is None else str(value).strip()
        out.append(record)
    wb.close()
    return out


def pick(row: dict[str, str], *keys: str) -> str:
    """Return the first non-empty value among the given possible header names."""
    for k in keys:
        v = row.get(k, "")
        if v:
            return v
    return ""
